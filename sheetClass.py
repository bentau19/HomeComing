from openpyxl import Workbook
import teacherClass
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.styles.colors import YELLOW
from openpyxl.styles.borders import Border, Side
from datetime import datetime
weekDays=["א","ב","ג","ד","ה","ו","ז"]
daysToNum ={"Sun": 0, "Mon": 1, "Tue": 2 ,"Wed": 3, "Thu": 4,"Fri":5,"Sat":6}
monthDays={"ינואר": 31,
"פבואר":28,
"מרץ": 31,
"אפריל": 30,
"מאי": 31,
"יוני": 30,
"יולי": 31,
"אוגוסט": 31,
"ספטמבר": 30,
"אוקטובר": 31,
"נובמבר": 30,
"דצמבר": 31
}
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
if(datetime.today().year%4==0):
    monthDays["פבואר"] = 29

class Sheet:
    Path = "template.xlsx"
    sheet = Workbook()

    sheet_obj = sheet.active
    month = "null"
    firstFlag={"row":5,"column":1} #5,1
    secondFlag = {"row": 5, "column": 1}
    teachers = []
    startDay=0


    def __init__(self):
        print("build template")
        self.sheet_setup()
        self.sheet_design(row=self.firstFlag["row"], column=self.firstFlag["column"],title="העדרויות חודש ")
        self.sheet_design(row=self.secondFlag["row"], column=self.secondFlag["column"], title="תוספות חודש ")
        self.sheet.save(str(datetime.today().year)+"-"+self.month+".xlsx")
        print("template build successfull")
    def sheet_setup(self):
        self.getData()
        self.secondFlag["row"]= self.firstFlag["row"]+6+len(self.teachers)


    def date_to_day(self,day):
        sum=((int(day)+self.startDay)%7)
        return sum
    def addHeaderCell(self,row,column,value):
        self.sheet_obj.cell(row = row, column = column).value=value
        self.sheet_obj.cell(row=row, column=column).fill = PatternFill(fgColor=YELLOW, fill_type = "solid")
        self.sheet_obj.cell(row=row, column=column).border = thin_border
    def sumDesign(self,row,column):
        self.sheet_obj.cell(row=row, column=column).value="סכום"
        self.sheet_obj.cell(row=row, column=column).fill = PatternFill(fgColor="ff99ff", fill_type="solid")
        self.sheet_obj.cell(row=row, column=column).border = thin_border
        for i in range(monthDays[self.month]):
            self.sheet_obj.cell(row=row, column=column+1+i).border = thin_border
            self.sheet_obj.cell(row=row, column=column+1+i).fill = PatternFill(fgColor="ff99ff", fill_type="solid")

    def teachers_list_design(self,row,column):
        for i in range(len(self.teachers)):
            self.addHeaderCell(row=row+i,column= column,value= self.teachers[i].name)
            for j in range(monthDays[self.month]):
                self.sheet_obj.cell(row=row+i, column=column+j+1).border = thin_border

    def dates_design(self,row,column):
        counter = self.startDay
        for i in range(monthDays[self.month]):
            self.sheet_obj.column_dimensions[get_column_letter((column+1+i))].width = 3
            self.addHeaderCell(row=row - 2,column= (column+1+i),value=  weekDays[counter])
            self.addHeaderCell(row=row - 1,column= (column+1+i),value= i+1)
            if(counter==6 or counter==5):
                for k in range(len(self.teachers)):
                    self.addHeaderCell(row=row+k, column=(column + 1 + i), value="")
            counter+=1
            if counter == 7:
                counter=0

    def add_num_to_place(self,teacher_location,date,num):
        number=int(num)
        int_date=int(date)
        int_location=int(teacher_location)
        self.sheet_obj.cell(row=self.firstFlag["row"]+teacher_location-1, column=self.firstFlag["column"]+int(int_date)).value = number
        try:
            current_sum_num = int(self.sheet_obj.cell(row=self.firstFlag["row"]+len(self.teachers), column=self.firstFlag["column"]+int_date).value)
        except:
            current_sum_num=0
        self.sheet_obj.cell(row=self.firstFlag["row"] + len(self.teachers) ,column=self.firstFlag["column"] + int_date).value=int(current_sum_num)+int(number)
        self.sheet.save(str(datetime.today().year)+"-"+self.month+".xlsx")
    def sheet_design(self,row,column,title):
        self.startDay = daysToNum[str(datetime.today().replace(day=1,month=self.monthNum()).ctime().split(" ")[0])]
        self.addHeaderCell(row=row-3,column=column,value=(title+self.month))
        self.sheet_obj.cell(row=row-3,column=column).font=Font(bold=True)
        self.sheet_obj.column_dimensions[get_column_letter(column)].width = 20
        self.addHeaderCell(row=row - 2,column= column, value="")
        self.addHeaderCell(row=row-1,column= column,value= "מורים/תאריך")
        self.sumDesign(row=row+len(self.teachers), column= column)
        self.dates_design(row,column)
        self.teachers_list_design(row,column)





    def getData(self):
        file = open('data.txt', encoding="utf8")
        with file as f:
            lines = f.readlines()
            self.month = lines[0].split(":")[1].lower().strip()
            lines.pop(0)
            for command in lines:
                data = command.split(":")
                teacher = teacherClass.Teacher(name=data[0], days=data[1].strip().split(","))
                self.teachers.append(teacher)

    def monthNum(self):
        return (list(monthDays).index(self.month)+1)